import csv
import json
import io
from datetime import date, datetime
from decimal import Decimal

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse

from .models import ProductionRecord, ColumnDefinition, ImportLog


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────

CORE_FIELDS = [
    'po_number', 'pallet_number', 'carton_number', 'pcb_label', 'chip_id',
    'serial_number',
    'stb_mac', 'bt_mac', 'wifi_ap_5g', 'wifi_ap_2g', 'wifi_client_5g', 'wifi_client_2g',
    'produce_dt'
]


def _get_all_headers():
    custom = list(ColumnDefinition.objects.filter(is_core=False).values_list('name', flat=True))
    return CORE_FIELDS + custom


def _record_to_row(record):
    row = {
        'po_number': record.po_number,
        'pallet_number': record.pallet_number,
        'carton_number': record.carton_number,
        'pcb_label': record.pcb_label,
        'chip_id': record.chip_id,
        'serial_number': record.serial_number,
        'stb_mac': record.stb_mac,
        'bt_mac': record.bt_mac,
        'wifi_ap_5g': record.wifi_ap_5g,
        'wifi_ap_2g': record.wifi_ap_2g,
        'wifi_client_5g': record.wifi_client_5g,
        'wifi_client_2g': record.wifi_client_2g,
        'produce_dt': str(record.produce_dt) if record.produce_dt else '',
    }
    row.update(record.extra_data)
    return row


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, (date, datetime)):
        return val.date() if isinstance(val, datetime) else val
    for fmt in (
        '%Y-%m-%d',
        '%d/%m/%Y',
        '%m/%d/%Y',
        '%d-%m-%Y',
        '%Y-%m-%d %H:%M',
        '%d/%m/%Y %H:%M',
        '%m/%d/%Y %H:%M',
        '%d-%m-%Y %H:%M',
    ):
        try:
            parsed = datetime.strptime(str(val).strip(), fmt)
            return parsed.date()
        except ValueError:
            continue
    raise ValueError(f'Cannot parse date: {val}')


def _import_rows(rows, user, skip_errors=True):
    log_errors = []
    success = 0
    custom_cols = {c.name: c for c in ColumnDefinition.objects.filter(is_core=False)}

    for i, row in enumerate(rows, start=2):
        try:
            serial = str(row.get('serial_number', '')).strip()
            if not serial:
                raise ValueError('serial_number is required')

            prod_date_raw = row.get('produce_dt', '')
            prod_date = _parse_date(prod_date_raw)
            if not prod_date:
                raise ValueError('produce_dt is required')

            extra = {}
            for col_name in custom_cols:
                if col_name in row:
                    extra[col_name] = str(row[col_name]) if row[col_name] is not None else ''

            record, created = ProductionRecord.objects.update_or_create(
                serial_number=serial,
                defaults={
                    'po_number': str(row.get('po_number', '') or ''),
                    'pallet_number': str(row.get('pallet_number', '') or ''),
                    'carton_number': str(row.get('carton_number', '') or ''),
                    'pcb_label': str(row.get('pcb_label', '') or ''),
                    'chip_id': str(row.get('chip_id', '') or ''),
                    'stb_mac': str(row.get('stb_mac', '') or ''),
                    'bt_mac': str(row.get('bt_mac', '') or ''),
                    'wifi_ap_5g': str(row.get('wifi_ap_5g', '') or ''),
                    'wifi_ap_2g': str(row.get('wifi_ap_2g', '') or ''),
                    'wifi_client_5g': str(row.get('wifi_client_5g', '') or ''),
                    'wifi_client_2g': str(row.get('wifi_client_2g', '') or ''),
                    'produce_dt': prod_date,
                    'extra_data': extra,
                    'updated_by': user,
                }
            )
            if created:
                record.created_by = user
                record.save(update_fields=['created_by'])
            success += 1
        except Exception as e:
            msg = f'Row {i}: {e}'
            log_errors.append(msg)
            if not skip_errors:
                raise

    return success, log_errors


# ─────────────────────────────────────────────
# IMPORT
# ─────────────────────────────────────────────

def import_csv(file_obj, user, skip_errors=True):
    decoded = file_obj.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(decoded))
    rows = list(reader)
    return _import_rows(rows, user, skip_errors)


def import_xlsx(file_obj, user, skip_errors=True):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        if any(v for v in row_dict.values()):
            rows.append(row_dict)
    return _import_rows(rows, user, skip_errors)


def import_json(file_obj, user, skip_errors=True):
    data = json.loads(file_obj.read().decode('utf-8'))
    if isinstance(data, dict):
        data = data.get('records', [data])
    return _import_rows(data, user, skip_errors)


# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

def export_csv(queryset):
    headers = _get_all_headers()
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="production_records.csv"'
    writer = csv.DictWriter(response, fieldnames=headers, extrasaction='ignore')
    writer.writeheader()
    for record in queryset:
        writer.writerow(_record_to_row(record))
    return response


def export_xlsx(queryset):
    headers = _get_all_headers()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Production Records'

    # Header styling
    header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=11)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(15, len(header) + 4)

    for row_idx, record in enumerate(queryset, 2):
        row = _record_to_row(record)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(header, ''))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="production_records.xlsx"'
    return response


def export_json(queryset):
    data = {'records': [_record_to_row(r) for r in queryset]}
    response = HttpResponse(
        json.dumps(data, indent=2, default=str),
        content_type='application/json'
    )
    return response


# ─────────────────────────────────────────────
# MAC-to-serial lookup import
# ─────────────────────────────────────────────

def _find_record_by_mac(mac):
    if not mac:
        return None
    mac_norm = str(mac).strip()
    if not mac_norm:
        return None
    return ProductionRecord.objects.filter(
        models.Q(stb_mac__iexact=mac_norm) |
        models.Q(bt_mac__iexact=mac_norm) |
        models.Q(wifi_ap_5g__iexact=mac_norm) |
        models.Q(wifi_ap_2g__iexact=mac_norm) |
        models.Q(wifi_client_5g__iexact=mac_norm) |
        models.Q(wifi_client_2g__iexact=mac_norm)
    ).first()


def _normalize_mac_header(headers):
    normalized = [h.strip().lower() for h in headers]
    for k in ('stb_mac', 'bt_mac', 'wifi_ap_5g', 'wifi_ap_2g', 'wifi_client_5g', 'wifi_client_2g', 'mac'):
        if k in normalized:
            return k
    return None


def map_macs_to_serial(rows):
    """rows: list of dicts containing at least one MAC value. Returns mapping list."""
    output = []
    for i, row in enumerate(rows, start=2):
        # try to find a mac field in row
        mac_value = ''
        for key in row:
            if key.strip().lower() in ('stb_mac', 'bt_mac', 'wifi_ap_5g', 'wifi_ap_2g', 'wifi_client_5g', 'wifi_client_2g', 'mac'):
                mac_value = str(row.get(key, '') or '').strip()
                if mac_value:
                    break
        if not mac_value:
            output.append({'row': i, 'mac': None, 'serial_number': None, 'error': 'No MAC value found'})
            continue

        rec = _find_record_by_mac(mac_value)
        if rec:
            output.append({'row': i, 'mac': mac_value, 'serial_number': rec.serial_number, 'status': 'found'})
        else:
            output.append({'row': i, 'mac': mac_value, 'serial_number': None, 'error': 'No matching record'})
    return output


def import_mac_list_csv(file_obj, skip_errors=False):
    decoded = file_obj.read().decode('utf-8-sig')
    reader = csv.DictReader(io.StringIO(decoded))
    rows = list(reader)
    return map_macs_to_serial(rows)


def import_mac_list_xlsx(file_obj, skip_errors=False):
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active
    headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # skip empty rows
        if not any(row):
            continue
        row_dict = {headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))}
        rows.append(row_dict)
    return map_macs_to_serial(rows)

    response['Content-Disposition'] = 'attachment; filename="production_records.json"'
    return response
